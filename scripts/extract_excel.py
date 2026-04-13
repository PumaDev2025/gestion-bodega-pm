"""
Script para extraer TODOS los datos del Excel
"Requerimientos Tendencias semana 06 de Abril 2026"
Rango: SEM16-2024 (Abril 2024) hasta SEM14-2026 (Abril 2026)
Incluye TODOS los movimientos historicos y TODOS los productos.
"""
import openpyxl
import re
from datetime import datetime, timedelta
from collections import defaultdict

EXCEL_PATH = "Requerimientos Tendencias semana 06 de Abril 2026.xlsx"
OUT_PRODUCTOS  = r"src\app\core\infrastructure\mock\productos.mock.ts"
OUT_CATEGORIAS = r"src\app\core\infrastructure\mock\categorias.mock.ts"
OUT_MOVIMIENTOS = r"src\app\core\infrastructure\mock\movimientos.mock.ts"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ─────────────────────────────────────────────
# 1. CLASIFICACION POR CODIGO MAXIMISE
# ─────────────────────────────────────────────
CATEGORIAS = {
    1:  {"nombre": "Herramientas Manuales",    "color": "#27ae60", "icono": "🔧"},
    2:  {"nombre": "Herramientas Eléctricas",  "color": "#e74c3c", "icono": "⚡"},
    3:  {"nombre": "Equipo de Medición",       "color": "#3498db", "icono": "📐"},
    4:  {"nombre": "Andamios y Estructuras",   "color": "#8e44ad", "icono": "🏗️"},
    5:  {"nombre": "Mobiliario de Obra",       "color": "#f39c12", "icono": "🪑"},
    6:  {"nombre": "Material Eléctrico",       "color": "#e67e22", "icono": "🔌"},
    7:  {"nombre": "EPP",                      "color": "#1abc9c", "icono": "🦺"},
    8:  {"nombre": "Equipos Electrónicos",     "color": "#2c3e50", "icono": "💻"},
    9:  {"nombre": "Soldadura y Corte",        "color": "#c0392b", "icono": "🔥"},
    10: {"nombre": "Medición y Topografía",    "color": "#16a085", "icono": "📏"},
}

CAT_DESC = {
    1:  "Alicates, llaves, martillos, niveles, huinchas, tizadores y herramientas de mano",
    2:  "Esmeriles angulares, taladros percutores, sierras circulares y rectificadores",
    3:  "Detectores de voltaje, medidores de gas, amperímetros y comprobadores",
    4:  "Verticales, horizontales, plataformas, rodapiés y accesorios de andamio Allround",
    5:  "Sillas, escritorios, cajoneras, lockers metálicos y muebles de faena",
    6:  "Extensiones eléctricas, tableros, ferrules aislados y marcadores de cable",
    7:  "Cascos, guantes, zapatos de seguridad, lentes, arnés y ropa de trabajo",
    8:  "Monitores, impresoras, radios portátiles y equipos electrónicos de obra",
    9:  "Soldadoras, electrodos, alimentadores MIG-MAG y hornos para soldadura",
    10: "Niveles topográficos, estaciones totales, prismas y equipos de nivelación",
}

def clasificar(cod, desc):
    cod  = (cod  or "").upper()
    desc = (desc or "").upper()

    # Andamios (36.xx)
    if ".36." in cod:
        return 4

    # Mobiliario (04.03, 12.02, 12.03)
    if any(x in cod for x in [".04.03.", ".12.02.", ".12.03.", ".12.04."]):
        return 5
    if any(x in desc for x in ["ESCRITORIO","SILLA","CAJONERA","LOCKER","MUEBLE","ESTANTE","PIZARRA"]):
        return 5

    # Soldadura / Corte (06.03, 06.04 cuando es corte)
    if ".06.03." in cod or any(x in desc for x in ["SOLDA","MIG","TIG","ELECTRODO","ARCO MANUAL","HORNO PARA SOLDA"]):
        return 9

    # Taladros / Herramientas eléctricas (06.02.10, 06.02.14, 08.08, 30.01.08.35, 30.01.08.33)
    if any(x in cod for x in [".06.02.10.",".06.02.14.",".08.08.",".08.06."]):
        return 2
    if any(x in desc for x in ["TALADRO","ESMERIL","RECTIFICADOR","SIERRA CALADORA","SIERRA CIRCULAR",
                                  "PISTOLA ANCLAJE","PISTOLA AIRE","HIDROLAVADORA","ROTOMARTILLO"]):
        return 2

    # Equipo de medición eléctrica (07.03, 08.07, 08.08.02.008 tipo comprobador)
    if any(x in cod for x in [".07.03.",".08.07.",".08.08.02."]):
        return 3
    if any(x in desc for x in ["DETECTOR DE VOLTAJE","MULTIMETRO","AMPERIMETRO","COMPROBADOR",
                                  "MEDIDOR HSL","TOXIRAE","TESTER","VOLTIMETRO"]):
        return 3

    # Topografía / medición mecánica (06.13)
    if ".06.13." in cod:
        return 10
    if any(x in desc for x in ["NIVEL TOPOGR","NIVEL ESFERIC","NIVEL OPTIC","ESTACION TOTAL",
                                  "MINI PRISMA","TOPOGRAFICO","MIRA TOPOGR","TEODOLITO"]):
        return 10

    # EPP (EPP, 08.03, zapato, casco, guante, arnés)
    if any(x in desc for x in ["CASCO","ZAPATO SEG","GUANTE","ARNES","LENTES SEG","PROTEC AUDITIVO",
                                  "CHALECO REFLEC","BOTIN","MASCARA","RESPIRADOR"]):
        return 7

    # Material eléctrico (07.01, ferrule, extension)
    if ".07.01." in cod or ".07.02." in cod:
        return 6
    if any(x in desc for x in ["EXTENSION ELECTR","TABLERO ELECTR","FERRULE","MARCADOR SHRINK",
                                  "CABLE ","DUCTO","BANDEJA","TOMACORRIENTE","INTERRUPTOR"]):
        return 6

    # Herramientas manuales (06.02.07, 06.02.08, 06.02.09, 06.04, 06.11, 30.01.08, etc.)
    if any(x in cod for x in [".06.02.07.",".06.02.08.",".06.02.09.",".06.04.",".06.11.",
                                ".30.01.",".32.13.",".01.01.",".06.08."]):
        return 1
    if any(x in desc for x in ["ALICATE","LLAVE ","MARTILLO","HUINCHA","NIVEL TORPEDO","TIZADOR",
                                  "COMBO","SERRUCHO","FORMON","CINCEL","PUNTA CENTRO","ESPATULA",
                                  "PLATACHO","LLANA","PLOMADA","ESCUADRA","PUNTO CENTRO","RASTRILLO",
                                  "BARRETILLA","UÑETA","DESTORNILLADOR","TECLE","CARRETILLA",
                                  "CAJA PARA HERR","MARCO SIERRA","TERRAJA","LLAVE DE GOLPE",
                                  "LLAVE AJUSTABLE","PIE DE METRO","DESVAINADOR"]):
        return 1

    # Equipos electrónicos
    if any(x in cod for x in ["P.02.","P.03.","S.P.03.","S.P.04.","P.01."]):
        return 8
    if any(x in desc for x in ["MONITOR","IMPRESORA","RADIO PORT","NOTEBOOK","PLASTIFICADORA",
                                  "ROTULADORA","TELEFONO"]):
        return 8

    # Default: herramientas manuales
    return 1


# ─────────────────────────────────────────────
# 2. EXTRAER PRODUCTOS (hoja Codigos = acumula 2024-2026)
# ─────────────────────────────────────────────
ws_cod = wb["Codigos"]
rows_cod = list(ws_cod.iter_rows(values_only=True))

# Columnas: 0=Descripcion, 1=COD.MAXIMISE, 2..8=proyectos, 9=Total general
PROYECTOS_COLS = {
    "P-4127": 2, "P-4132": 3, "P-4166": 4,
    "P-4219": 5, "P-4231": 6, "P-4257": 7, "P-4254": 8
}

productos_raw = {}
for row in rows_cod[3:]:
    desc  = str(row[0]).strip() if row[0] else None
    cod   = str(row[1]).strip() if row[1] and str(row[1]).strip() not in ("(en blanco)","None","") else None
    total = row[9] if isinstance(row[9], (int, float)) else 0
    if not desc or desc in ("Descripcion","Total general"):
        continue
    key = cod if cod else f"_DESC_{desc[:50]}"
    if key not in productos_raw:
        productos_raw[key] = {"cod": cod, "desc": desc, "total": 0}
        for pname in PROYECTOS_COLS:
            productos_raw[key][pname.replace("-","").lower()] = 0
    if cod and not productos_raw[key]["cod"]:
        productos_raw[key]["cod"] = cod
    if len(desc) < len(productos_raw[key]["desc"]):
        productos_raw[key]["desc"] = desc
    productos_raw[key]["total"] += total
    for pname, col in PROYECTOS_COLS.items():
        val = row[col] if isinstance(row[col], (int, float)) else 0
        productos_raw[key][pname.replace("-","").lower()] += val

productos_list = [v for v in productos_raw.values() if v["total"] > 0]
productos_list.sort(key=lambda x: -x["total"])
print(f"Total productos únicos con demanda histórica > 0: {len(productos_list)}")

# ─────────────────────────────────────────────
# 2b. PRE-CALCULAR STOCK REAL DESDE BD2
#     (antes de generar productos mock)
# ─────────────────────────────────────────────
ws_bd2   = wb["BD2"]
rows_bd2 = list(ws_bd2.iter_rows(values_only=True))
print(f"Total filas BD2 (sin header): {len(rows_bd2)-1}")

# Por cada código: lista de cantidades despachadas y máximo solicitado por proyecto
stock_stats_bd2 = {}
for _r in rows_bd2[1:]:
    _cod  = str(_r[8]).strip() if _r[8] and str(_r[8]) not in ("None","") else None
    if not _cod:
        continue
    _cod  = _cod.upper()
    _desp = _r[17] if isinstance(_r[17], (int, float)) and _r[17] else 0
    _sol  = _r[5]  if isinstance(_r[5],  (int, float)) and _r[5]  else 0
    _proj = str(_r[1]).strip() if _r[1] else "N/A"
    if _cod not in stock_stats_bd2:
        stock_stats_bd2[_cod] = {"desps": [], "max_sol_proy": {}}
    if _desp > 0:
        stock_stats_bd2[_cod]["desps"].append(_desp)
    if _sol > 0:
        _prev = stock_stats_bd2[_cod]["max_sol_proy"].get(_proj, 0)
        stock_stats_bd2[_cod]["max_sol_proy"][_proj] = max(_prev, _sol)

def calc_stock(cod, n_projs_activos, total_pivot):
    """Calcula stockActual, stockMinimo, stockMaximo realistas.
    Usa despachos reales de BD2 (por código) en lugar de sumar todos los proyectos."""
    cod_up = (cod or "").upper()
    st = stock_stats_bd2.get(cod_up)
    if st and st["desps"]:
        # Stock basado en el mayor despacho individual (lo que el almacén envía de una vez)
        max_desp     = max(st["desps"])
        avg_desp     = sum(st["desps"]) / len(st["desps"])
        stock_actual = max(1, round(max_desp * 1.2))          # buffer 20% sobre max despacho
        stock_min    = max(1, round(avg_desp))                # promedio de despachos como mínimo
        stock_max    = max(stock_actual + 1, round(max_desp * 3))
    else:
        # Sin despachos reales: dividir pivot entre proyectos activos
        por_proy     = max(1, round(total_pivot / max(1, n_projs_activos)))
        stock_actual = por_proy
        stock_min    = max(1, por_proy // 2)
        stock_max    = max(stock_actual + 1, por_proy * 3)
    return stock_actual, stock_min, stock_max

cat_count = defaultdict(int)
for p in productos_list:
    cat = clasificar(p["cod"], p["desc"])
    p["catId"] = cat
    cat_count[cat] += 1
for cat_id, cnt in sorted(cat_count.items()):
    print(f"  Categoría {cat_id} ({CATEGORIAS[cat_id]['nombre']}): {cnt} productos")

# ─────────────────────────────────────────────
# 3. GENERAR CATEGORIAS MOCK
# ─────────────────────────────────────────────
lines_cat = [
    "import { Categoria } from '../../domain/models';\n",
    "\n",
    "// Categorías generadas desde Requerimientos Tendencias SEM14 - 06 Abril 2026\n",
    "export const MOCK_CATEGORIAS: Categoria[] = [\n",
]
for cat_id, info in CATEGORIAS.items():
    cnt = cat_count.get(cat_id, 0)
    lines_cat.append(f"  {{\n")
    lines_cat.append(f"    id: {cat_id}, nombre: '{info['nombre']}',\n")
    desc_map = {
        1:  "Alicates, llaves, martillos, niveles, huinchas, tizadores y herramientas de mano",
        2:  "Esmeriles angulares, taladros percutores, sierras circulares y rectificadores",
        3:  "Detectores de voltaje, medidores de gas, amperímetros y comprobadores",
        4:  "Verticales, horizontales, plataformas, rodapiés y accesorios de andamio Allround",
        5:  "Sillas, escritorios, cajoneras, lockers metálicos y muebles de faena",
        6:  "Extensiones eléctricas, tableros, ferrules aislados y marcadores de cable",
        7:  "Cascos, guantes, zapatos de seguridad, lentes, arnés y ropa de trabajo",
        8:  "Monitores, impresoras, radios portátiles y equipos electrónicos de obra",
        9:  "Soldadoras, electrodos, alimentadores MIG-MAG y hornos para soldadura",
        10: "Niveles topográficos, estaciones totales, prismas y equipos de nivelación",
    }
    lines_cat.append(f"    descripcion: '{desc_map[cat_id]}',\n")
    lines_cat.append(f"    color: '{info['color']}', icono: '{info['icono']}', cantidadProductos: {cnt}\n")
    lines_cat.append(f"  }},\n")
lines_cat.append("];\n")

with open(OUT_CATEGORIAS, "w", encoding="utf-8") as f:
    f.writelines(lines_cat)
print(f"\nArchivo generado: {OUT_CATEGORIAS}")


# ─────────────────────────────────────────────
# 4. GENERAR productos.mock.ts
# ─────────────────────────────────────────────
def gen_ubicacion(cat_id, seq):
    prefijos = {1:"A",2:"B",3:"C",4:"D",5:"E",6:"F",7:"G",8:"H",9:"I",10:"J"}
    p = prefijos.get(cat_id, "Z")
    estante  = (seq - 1) // 5 + 1
    posicion = (seq - 1) % 5 + 1
    return f"{p}-{estante:02d}-{posicion:02d}"

def gen_precio(cat_id, total):
    import random
    random.seed(total + cat_id)
    base = {1:15000,2:75000,3:45000,4:25000,5:65000,6:12000,7:18000,8:95000,9:55000,10:85000}
    return int(base.get(cat_id,15000) * random.uniform(0.6,1.4) // 100 * 100)

def gen_unidad(desc):
    d = (desc or "").upper()
    if any(x in d for x in ["FERRULE","MARCADOR SHRINK"]): return "unidades"
    if "CABLE " in d: return "metros"
    if "M3" in d or "M\u00b3" in d: return "m\u00b3"
    return "unidades"

def clean_str(s):
    return (str(s) if s else "").replace("'", "\\'").replace("\n", " ").strip()

lines_prod = [
    "import { Producto } from '../../domain/models';\n\n",
    f"// Inventario completo generado desde Requerimientos Tendencias 2024-2026\n",
    f"// Total: {len(productos_list)} productos con demanda histórica registrada\n\n",
]

PROD_CHUNK = 200
prod_items = []
cat_seq = defaultdict(int)
for idx, p in enumerate(productos_list, start=1):
    cat_id = p["catId"]
    cat_seq[cat_id] += 1
    codigo  = clean_str(p["cod"]) if p["cod"] else f"SIN-COD-{idx:04d}"
    nombre  = clean_str(p["desc"].title())
    desc_ts = clean_str(f'{p["desc"].title()} \u2014 demanda hist\u00f3rica acumulada: {int(p["total"])} un')
    ubic    = gen_ubicacion(cat_id, cat_seq[cat_id])
    n_projs_activos = sum(1 for pn in PROYECTOS_COLS if p.get(pn.replace("-","").lower(), 0) > 0)
    stock_actual, stock_min, stock_max = calc_stock(p["cod"], n_projs_activos, int(p["total"]))
    precio       = gen_precio(cat_id, int(p["total"]))
    unidad       = gen_unidad(p["desc"])
    # Lista de proyectos que solicitaron este producto (de la hoja Codigos)
    proy_list = [pn for pn in PROYECTOS_COLS if p.get(pn.replace("-","").lower(), 0) > 0]
    proy_ts   = ", ".join(f"'{pn}'" for pn in proy_list)
    prod_items.append((idx, cat_id, codigo, nombre, desc_ts, ubic,
                       stock_actual, stock_min, stock_max, unidad, precio, proy_ts))

n_prod_chunks = (len(prod_items) + PROD_CHUNK - 1) // PROD_CHUNK
prod_chunk_names = [f"_PROD_CHUNK_{i+1}" for i in range(n_prod_chunks)]

for ci, cname in enumerate(prod_chunk_names):
    lines_prod.append(f"const {cname}: Producto[] = [\n")
    for (idx, cat_id, codigo, nombre, desc_ts, ubic,
         stock_actual, stock_min, stock_max, unidad, precio, proy_ts) in prod_items[ci*PROD_CHUNK:(ci+1)*PROD_CHUNK]:
        lines_prod += [
            f"  {{\n",
            f"    id: {idx}, codigo: '{codigo}', nombre: '{nombre}',\n",
            f"    descripcion: '{desc_ts}', categoriaId: {cat_id},\n",
            f"    categoriaNombre: '{CATEGORIAS[cat_id]['nombre']}', ubicacion: '{ubic}', stockActual: {stock_actual},\n",
            f"    stockMinimo: {stock_min}, stockMaximo: {stock_max}, unidadMedida: '{unidad}',\n",
            f"    precioUnitario: {precio}, estado: 'activo',\n",
            f"    fechaIngreso: new Date('2024-04-15'), ultimaActualizacion: new Date('2026-04-10'),\n",
            f"    proyectos: [{proy_ts}]\n",
            f"  }},\n",
        ]
    lines_prod.append("];\n\n")

prod_chunks_concat = ", ...".join(prod_chunk_names)
lines_prod.append(f"export const MOCK_PRODUCTOS: Producto[] = [...{prod_chunks_concat}];\n")

with open(OUT_PRODUCTOS, "w", encoding="utf-8") as f:
    f.writelines(lines_prod)
print(f"Generado: {OUT_PRODUCTOS} ({len(productos_list)} productos)")


# ─────────────────────────────────────────────
# 5. PARSEAR FECHA DESDE CAMPO SEM
# ─────────────────────────────────────────────
def parse_fecha_sem(sem_str, fallback_fecha=None):
    """Parsear fecha de inicio de semana desde el campo SEM.
    Formatos:
      'SEM14 06/04/2026 Al 12/04/2026'
      'SEM16-2024 15/04/2024 Al 21/04/2024'
      'SEM01 29/12/2025 Al 04/01/2026'
      'Pendientes'
    """
    if not sem_str:
        return fallback_fecha or datetime(2026, 4, 7)
    match = re.search(r'(\d{2}/\d{2}/\d{4})', str(sem_str))
    if match:
        try:
            return datetime.strptime(match.group(1), "%d/%m/%Y")
        except:
            pass
    return fallback_fecha or datetime(2026, 4, 7)

# ─────────────────────────────────────────────
# 6. EXTRAER MOVIMIENTOS COMPLETOS BD2 (2024-2026)
#    rows_bd2 ya cargado en sección 2b
# ─────────────────────────────────────────────

# Mapa codigo MAXIMISE → id producto
cod_to_id = {}
for idx, p in enumerate(productos_list, start=1):
    if p["cod"]:
        cod_to_id[p["cod"].strip().upper()] = idx

def buscar_prod_id(cod_excel, desc_excel):
    if cod_excel:
        pid = cod_to_id.get(cod_excel.strip().upper())
        if pid:
            return pid
    if desc_excel:
        needle = desc_excel.strip().upper()[:20]
        for p in productos_list:
            if needle in p["desc"].upper():
                return productos_list.index(p) + 1
    return 1

def to_num(v):
    return v if isinstance(v, (int, float)) and v else 0

movimientos = []
skipped = 0
for r in rows_bd2[1:]:
    cod_excel  = str(r[8]).strip() if r[8] and str(r[8]) not in ("None","") else None
    desc_excel = str(r[9]).strip() if r[9] else "SIN DESCRIPCION"
    cant_desp  = to_num(r[17])
    cant_sol   = to_num(r[5])
    proyecto   = str(r[1]).strip() if r[1] else "N/A"
    estado     = str(r[19]).strip() if r[19] else ""
    guia       = str(r[21]).strip() if r[21] and str(r[21]) not in ("None","") else None
    obs_raw    = str(r[24]).strip() if r[24] and str(r[24]) not in ("None","") else ""
    prioridad  = str(r[7]).strip() if r[7] else ""
    sem_desp   = str(r[27]).strip() if r[27] else ""
    sem_rec    = str(r[4]).strip()  if r[4]  else ""
    fecha_raw  = r[20]

    if cant_desp > 0:
        tipo     = "salida"
        cantidad = int(cant_desp)
        if hasattr(fecha_raw, "strftime"):
            fecha_mov = fecha_raw
        else:
            fecha_mov = parse_fecha_sem(sem_desp, parse_fecha_sem(sem_rec))
    elif cant_sol > 0:
        tipo     = "salida"
        cantidad = int(cant_sol)
        fecha_mov = parse_fecha_sem(sem_rec)
    else:
        skipped += 1
        continue

    movimientos.append({
        "tipo":       tipo,
        "productoId": buscar_prod_id(cod_excel, desc_excel),
        "prodNombre": desc_excel,
        "prodCodigo": cod_excel,
        "cantidad":   cantidad,
        "fecha":      fecha_mov,
        "proyecto":   proyecto,
        "prioridad":  prioridad,
        "guia":       guia,
        "obs":        obs_raw[:120],
        "estado":     estado,
        "semDesp":    sem_desp,
    })

movimientos.sort(key=lambda x: x["fecha"])
print(f"Movimientos válidos: {len(movimientos)} | Omitidos (sin cantidad): {skipped}")

# ─────────────────────────────────────────────
# 7. GENERAR movimientos.mock.ts
#    Se divide en chunks de 500 para evitar TS2590
#    "Expression produces a union type too complex"
# ─────────────────────────────────────────────
CHUNK_SIZE = 500

def render_mov_item(mov_id, m):
    fecha_str   = m["fecha"].strftime("%Y-%m-%dT08:00:00") if hasattr(m["fecha"],"strftime") else "2026-04-07T08:00:00"
    prod_nombre = clean_str(m["prodNombre"].title())
    motivo      = clean_str(f'Despacho {m["proyecto"]}')
    obs_final   = clean_str(m["obs"]) if m["obs"] else clean_str(f'Estado: {m["estado"]} | {m["semDesp"]}')
    lines = [f"  {{\n", f"    id: {mov_id}, tipo: '{m['tipo']}', productoId: {m['productoId']},\n"]
    lines.append(f"    productoNombre: '{prod_nombre}',\n")
    if m["prodCodigo"]:
        lines.append(f"    productoCodigo: '{clean_str(m['prodCodigo'])}',\n")
    lines += [
        f"    cantidad: {m['cantidad']}, fecha: new Date('{fecha_str}'),\n",
        f"    responsable: 'Bodega Central P-1010',\n",
        f"    motivo: '{motivo}',\n",
    ]
    if m["guia"]:
        lines.append(f"    documento: '{clean_str(m['guia'])}',\n")
    lines.append(f"    observaciones: '{obs_final[:120]}',\n")
    lines.append(f"    proyecto: '{clean_str(m['proyecto'])}'\n")
    lines.append(f"  }},\n")
    return lines

# Calcular número de chunks
n_chunks = (len(movimientos) + CHUNK_SIZE - 1) // CHUNK_SIZE
chunk_names = [f"_MOV_CHUNK_{i+1}" for i in range(n_chunks)]

lines_mov = [
    "import { Movimiento } from '../../domain/models';\n\n",
    f"// Historial completo de movimientos 2024-2026 ({len(movimientos)} registros)\n",
    f"// Dividido en {n_chunks} chunks de {CHUNK_SIZE} para evitar TS2590\n\n",
]

# Generar cada chunk como const tipada
for ci, chunk_name in enumerate(chunk_names):
    start_idx = ci * CHUNK_SIZE
    end_idx   = min(start_idx + CHUNK_SIZE, len(movimientos))
    lines_mov.append(f"const {chunk_name}: Movimiento[] = [\n")
    for mov_id, m in enumerate(movimientos[start_idx:end_idx], start=start_idx + 1):
        lines_mov.extend(render_mov_item(mov_id, m))
    lines_mov.append("];\n\n")

# Export final que concatena todos los chunks
chunks_concat = ", ...".join(chunk_names)
lines_mov.append(f"export const MOCK_MOVIMIENTOS: Movimiento[] = [...{chunks_concat}];\n")

with open(OUT_MOVIMIENTOS, "w", encoding="utf-8") as f:
    f.writelines(lines_mov)
print(f"Generado: {OUT_MOVIMIENTOS} ({len(movimientos)} movimientos en {n_chunks} chunks)")
print("\n\u2713 Done \u2014 historial completo 2024-2026 generado.")
